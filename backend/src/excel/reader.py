import openpyxl


def read_first_column(file_name):
    sheet = load_first_sheet(file_name)
    first_column = list(sheet.columns)[0]

    values = []

    for cell in first_column:
        if cell.value is None:
            break
        values.append(cell.value)

    return values


def load_first_sheet(file_name):
    workbook = load_workbook(file_name)
    first_sheet = workbook.sheetnames[0]
    return workbook[first_sheet]


def load_workbook(file_name):
    return openpyxl.load_workbook(file_name)
