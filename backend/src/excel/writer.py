from openpyxl.styles import PatternFill
import openpyxl


def workbook_path():
    return workbook_parent_dir() + workbook_name()


def workbook_name():
    return 'coffee_roulette_pairings.xlsx'


def workbook_parent_dir():
    return './resources/'


def create_workbook():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Pairings'

    basket = ['apples', 'bananas', 'blueberries', 'cherries']
    populate_matrix_row_header(sheet, basket)
    populate_matrix_column_header(sheet, basket)

    num_rows = len(basket)
    fill_diagonal(sheet, num_rows)

    path = workbook_path()
    workbook.save(path)


def populate_matrix_row_header(sheet, data):
    for index, value in enumerate(data):
        # +2 because we want the first cell to be A2, not A0
        cell = 'A' + str(index + 2)
        sheet[cell] = value


def populate_matrix_column_header(sheet, data):
    for index, value in enumerate(data):
        # +2 because we want the first cell to be B1, not B0
        col_num = index + 2
        cell = sheet.cell(row=1, column=col_num)
        cell = cell.column_letter + str(cell.row)
        sheet[cell] = value


# A person cannot be paired with themselves, so we block this out with a solid color
def fill_diagonal(sheet, num_rows):
    for i in range(num_rows):
        # First cell: B2
        cell = sheet.cell(row=(i + 2), column=(i + 2))
        cell = cell.column_letter + str(cell.row)
        sheet[cell].fill = PatternFill("solid", fgColor="000000")